from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from prj_strategic_pricing.calculator import PricingResult


INPUT_SHEETS = (
    "deal_brief",
    "scope",
    "wbs",
    "rates",
    "non_labor_costs",
    "value_market_anchor",
    "strategy",
    "approval",
)


def read_input_workbook(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True)
    return {
        sheet_name: _sheet_to_records(workbook[sheet_name])
        for sheet_name in INPUT_SHEETS
        if sheet_name in workbook.sheetnames
    }


def write_result_workbook(
    path: str | Path,
    input_data: dict[str, list[dict[str, Any]]],
    result: PricingResult,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_rows(
        workbook,
        "customer_quotation",
        ["item", "amount"],
        [
            ["Labor Cost", result.labor_cost],
            ["Non-Labor Cost", result.non_labor_cost],
            ["Risk Buffer", result.risk_buffer],
            ["Minimum Margin", result.minimum_margin],
            ["Sub Total", result.proposed_price],
            ["VAT", result.proposed_price * 0.10],
            ["Total Including VAT", result.proposed_price * 1.10],
            ["Walk-away Price", result.walkaway_price],
        ],
    )
    _write_rows(
        workbook,
        "cost_breakdown",
        ["category", "amount"],
        [
            ["Labor Cost", result.labor_cost],
            ["Non-Labor Cost", result.non_labor_cost],
            ["Risk Buffer", result.risk_buffer],
            ["Minimum Margin", result.minimum_margin],
            ["Price Cost", result.price_cost],
        ],
    )
    _write_records(workbook, "schedule", input_data.get("wbs", []))
    _write_package_options(workbook, result, input_data.get("strategy", []))
    _write_rows(
        workbook,
        "pricing_memo",
        ["candidate", "amount", "meaning"],
        [
            ["Price Ceiling", result.price_ceiling, "Customer saving/value x ratio"],
            ["Price Market", result.price_market, "Competitor average x ratio"],
            ["Price Anchor", result.price_anchor, "Annual anchor x project months / 12"],
            ["Price Cost", result.price_cost, "Labor + Non-Labor + Buffer + Margin"],
            ["Proposed Price", result.proposed_price, "Selected quote amount"],
            ["Required M/M", result.required_mm, "WBS based effort"],
            ["Available M/M", result.available_mm or "", "Budget based effort capacity"],
            ["M/M Feasible", str(result.mm_is_feasible), "Required M/M <= Available M/M"],
        ],
    )
    _write_rows(
        workbook,
        "checklist",
        ["check", "status"],
        [[item, "Review"] for item in _checklist()] + [[warning, "Warning"] for warning in result.warnings],
    )
    _write_records(workbook, "assumptions_scope", input_data.get("scope", []))
    _write_records(workbook, "approval_status", input_data.get("approval", []))

    workbook.save(path)


def create_template_workbook(path: str | Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_rows(
        workbook,
        "deal_brief",
        ["key", "value"],
        [
            ["client", "Example Client"],
            ["deal_type", "Build"],
            ["project_months", 12],
            ["budget_status", "Confirmed"],
            ["competition", "Competitive bid"],
        ],
    )
    _write_rows(
        workbook,
        "scope",
        ["category", "in_scope", "out_of_scope", "assumption"],
        [
            ["Data", "Excel/PDF estimate data for 12 months", "Missing data recovery", "Client provides source files"],
            ["Function", "Quotation automation and dashboard", "Expansion to other departments", "Change requests are re-estimated"],
            ["Operation", "Handover and basic defect support", "24/7 managed operation", "UAT completes within agreed window"],
        ],
    )
    _write_rows(
        workbook,
        "wbs",
        ["phase", "workstream", "task", "role", "grade", "mm", "basis"],
        [
            ["Discovery", "PMO", "Requirements workshop", "PM", "Senior", 1.5, "Stakeholders and workshops"],
            ["Build", "Data", "Data mapping and cleansing", "Data Engineer", "Mid", 4.0, "Source count and quality"],
            ["Build", "AI", "RAG/model workflow", "AI Engineer", "Senior", 5.0, "KPI and model complexity"],
            ["Build", "Backend", "API and DB", "Backend Engineer", "Mid", 3.5, "API and integration count"],
            ["Test", "QA/UAT", "Test and defect fix", "QA", "Mid", 2.0, "Test cases"],
        ],
    )
    _write_rows(
        workbook,
        "rates",
        ["grade", "monthly_rate"],
        [["Executive", 18_000_000], ["Senior", 14_000_000], ["Mid", 10_000_000], ["Junior", 7_000_000]],
    )
    _write_rows(
        workbook,
        "non_labor_costs",
        ["category", "item", "amount"],
        [
            ["GPU", "Training/inference environment", 30_000_000],
            ["API", "LLM/OCR API usage", 15_000_000],
            ["DB", "Vector DB and storage", 10_000_000],
            ["License", "Third-party component", 20_000_000],
            ["Education", "Training materials and sessions", 5_000_000],
        ],
    )
    _write_rows(
        workbook,
        "value_market_anchor",
        ["key", "value"],
        [
            ["customer_saving", 1_920_000_000],
            ["ceiling_ratio", 0.60],
            ["competitor_price_1", 1_200_000_000],
            ["competitor_price_2", 1_100_000_000],
            ["market_ratio", 0.80],
            ["anchor_annual", 1_000_000_000],
            ["project_months", 12],
        ],
    )
    _write_rows(
        workbook,
        "strategy",
        ["key", "value"],
        [
            ["risk_buffer_rate", 0.10],
            ["minimum_margin_rate", 0.20],
            ["walkaway_multiplier", 1.00],
            ["core_multiplier", 0.75],
            ["standard_multiplier", 1.00],
            ["premium_multiplier", 1.25],
            ["payment_terms", "30% upfront, 40% milestone, 30% UAT"],
            ["discount_condition", "Scope reduction, upfront payment, or reference rights required"],
        ],
    )
    _write_rows(
        workbook,
        "approval",
        ["team", "owner", "status", "notes"],
        [
            ["PM", "", "Pending", "Review WBS and schedule M/M"],
            ["Tech", "", "Pending", "Review architecture and delivery risk"],
            ["Infra", "", "Pending", "Review GPU/API/DB cost"],
            ["Finance", "", "Pending", "Review rates, margin, and cash flow"],
            ["Sales", "", "Pending", "Review market and negotiation room"],
            ["Legal", "", "Pending", "Review quote terms, IP, data, and liability"],
        ],
    )
    workbook.save(path)


def _sheet_to_records(sheet: Any) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records = []
    for row in rows[1:]:
        record = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def _write_records(workbook: Workbook, sheet_name: str, records: list[dict[str, Any]]) -> None:
    if not records:
        _write_rows(workbook, sheet_name, ["message"], [["No data"]])
        return
    headers = list(records[0].keys())
    rows = [[record.get(header, "") for header in headers] for record in records]
    _write_rows(workbook, sheet_name, headers, rows)


def _write_rows(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _format_sheet(sheet)


def _write_package_options(workbook: Workbook, result: PricingResult, strategy_rows: list[dict[str, Any]]) -> None:
    strategy = {str(row.get("key", "")).strip(): row.get("value") for row in strategy_rows}
    rows = []
    for name, key in (("Core", "core_multiplier"), ("Standard", "standard_multiplier"), ("Premium", "premium_multiplier")):
        multiplier = float(strategy.get(key) or 1.0)
        rows.append([name, result.proposed_price * multiplier, _package_scope(name)])
    _write_rows(workbook, "package_options", ["package", "price", "scope"], rows)


def _package_scope(name: str) -> str:
    scopes = {
        "Core": "Core functions, limited data/users, basic report",
        "Standard": "Main functions, main data, dashboard, UAT, education",
        "Premium": "Multi-department expansion, advanced automation, monitoring, extra training",
    }
    return scopes[name]


def _format_sheet(sheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def _checklist() -> list[str]:
    return [
        "Go/No-go 5 Lens reviewed",
        "Customer saving/value and 60% ceiling calculated",
        "Competitor or alternative price basis captured",
        "12-month 1B KRW anchor converted by duration",
        "Required M/M estimated from WBS",
        "Schedule M/M equals breakdown M/M",
        "GPU/API/DB/OCR/license costs separated from labor",
        "Risk buffer and minimum margin included",
        "Core/Standard/Premium options reviewed",
        "Discount tied to scope, terms, reference, or payment condition",
        "Out-of-Scope and Change Request terms stated",
        "Finance/PM/Sales/Infra/Legal/Approver reviews completed",
    ]

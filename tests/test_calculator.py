import tempfile
import unittest
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from strategic_pricing_calculator.calculator import calculate_pricing
from strategic_pricing_calculator.workbook import (
    create_template_workbook,
    read_input_workbook,
    write_result_workbook,
)


class PricingCalculatorTest(unittest.TestCase):
    def test_calculates_four_candidate_prices(self):
        data = {
            "wbs": [
                {"grade": "Senior", "mm": 2},
                {"grade": "Mid", "mm": 3},
            ],
            "rates": [
                {"grade": "Senior", "monthly_rate": 10_000_000},
                {"grade": "Mid", "monthly_rate": 5_000_000},
            ],
            "non_labor_costs": [{"amount": 15_000_000}],
            "value_market_anchor": [
                {"key": "customer_saving", "value": 200_000_000},
                {"key": "competitor_price_1", "value": 150_000_000},
                {"key": "project_months", "value": 6},
            ],
            "strategy": [
                {"key": "risk_buffer_rate", "value": 0.1},
                {"key": "minimum_margin_rate", "value": 0.2},
            ],
        }

        result = calculate_pricing(data)

        self.assertEqual(result.labor_cost, 35_000_000)
        self.assertEqual(result.non_labor_cost, 15_000_000)
        self.assertEqual(result.price_ceiling, 120_000_000)
        self.assertEqual(result.price_market, 120_000_000)
        self.assertEqual(result.price_anchor, 500_000_000)
        self.assertGreater(result.price_cost, 0)
        self.assertGreaterEqual(result.proposed_price, result.price_cost)

    def test_template_can_round_trip_to_result_workbook(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"

            create_template_workbook(input_path)
            data = read_input_workbook(input_path)
            result = calculate_pricing(data)
            write_result_workbook(output_path, data, result)

            workbook = load_workbook(output_path, data_only=True)
            self.assertIn("customer_quotation", workbook.sheetnames)
            self.assertIn("pricing_memo", workbook.sheetnames)
            self.assertIn("package_options", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
